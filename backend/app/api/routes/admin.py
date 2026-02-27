from fastapi import APIRouter, Depends, Header, HTTPException, UploadFile, File, Form
from fastapi.encoders import jsonable_encoder
from sqlalchemy.orm import Session
import requests
import csv
from io import BytesIO, StringIO
from openpyxl import load_workbook

from db.database import SessionLocal
from db.models import Message, Feedback, AdminUser, TrustedSender, BlockedEntity, BlockEvent, AppSetting
from core.threat_rules import extract_domain_like
from collections import defaultdict
from datetime import datetime, timedelta
from core.auth import verify_token
from core.retrain_guard import can_retrain
from core.retraining import trigger_retrain
import os

LAST_TRAINING_SUMMARY = {}

def require_admin(authorization: str | None = Header(default=None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1]
    payload = verify_token(token)
    email = payload.get("email")
    if not email:
        raise HTTPException(status_code=401, detail="Invalid token")
    db = SessionLocal()
    try:
        user = db.query(AdminUser).filter(AdminUser.email == email).first()
        if not user or user.is_active != 1:
            raise HTTPException(status_code=403, detail="User disabled")
    finally:
        db.close()
    return email


router = APIRouter(
    prefix="/admin",
    tags=["Admin"],
    dependencies=[Depends(require_admin)]
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/messages")
def list_messages(db: Session = Depends(get_db)):
    items = (
        db.query(Message)
        .order_by(Message.created_at.desc())
        .limit(100)
        .all()
    )
    return jsonable_encoder(items)

@router.get("/feedback")
def list_feedback(db: Session = Depends(get_db)):
    items = (
        db.query(Feedback)
        .order_by(Feedback.created_at.desc())
        .limit(100)
        .all()
    )
    return jsonable_encoder(items)

@router.get("/stats")
def stats(db: Session = Depends(get_db)):
    total = db.query(Feedback).count()
    phishing = db.query(Feedback).filter(Feedback.is_phishing == 1).count()
    legit = total - phishing
    total_messages = db.query(Message).count()
    blocked = db.query(Message).filter(Message.verdict == "BLOCK").count()
    flagged = db.query(Message).filter(Message.verdict == "FLAG").count()
    allowed = total_messages - blocked - flagged
    return {
        "total_feedback": total,
        "phishing": phishing,
        "legit": legit,
        "total_messages": total_messages,
        "allowed": allowed,
        "blocked": blocked,
        "flagged": flagged
    }


@router.get("/blocked-summary")
def blocked_summary(db: Session = Depends(get_db)):
    events = db.query(BlockEvent).all()
    by_reason = {}
    by_source = {}
    for e in events:
        reason = e.reason or "unknown"
        source = e.source or "unknown"
        by_reason[reason] = by_reason.get(reason, 0) + 1
        by_source[source] = by_source.get(source, 0) + 1
    return {"by_reason": by_reason, "by_source": by_source}


@router.get("/analytics")
def analytics(db: Session = Depends(get_db)):
    # Last 7 days trend
    since = datetime.utcnow() - timedelta(days=7)
    messages = db.query(Message).filter(Message.created_at >= since).all()
    feedback = db.query(Feedback).filter(Feedback.created_at >= since).all()
    events = db.query(BlockEvent).filter(BlockEvent.created_at >= since).all()

    day_key = lambda dt: dt.strftime("%Y-%m-%d")

    verdict_trend = defaultdict(lambda: {"ALLOW": 0, "FLAG": 0, "BLOCK": 0})
    for m in messages:
        verdict_trend[day_key(m.created_at)][m.verdict or "ALLOW"] += 1

    feedback_trend = defaultdict(lambda: {"phishing": 0, "legit": 0})
    for f in feedback:
        label = "phishing" if f.is_phishing == 1 else "legit"
        feedback_trend[day_key(f.created_at)][label] += 1

    block_reasons = defaultdict(int)
    block_sources = defaultdict(int)
    for e in events:
        block_reasons[e.reason or "unknown"] += 1
        block_sources[e.source or "unknown"] += 1

    # Risk score histogram buckets
    buckets = {"0-0.2": 0, "0.2-0.4": 0, "0.4-0.6": 0, "0.6-0.8": 0, "0.8-1.0": 0}
    for m in messages:
        score = m.risk_score or 0
        if score < 0.2:
            buckets["0-0.2"] += 1
        elif score < 0.4:
            buckets["0.2-0.4"] += 1
        elif score < 0.6:
            buckets["0.4-0.6"] += 1
        elif score < 0.8:
            buckets["0.6-0.8"] += 1
        else:
            buckets["0.8-1.0"] += 1

    # TLD distribution (from domain-like tokens)
    tld_counts = defaultdict(int)
    for m in messages:
        for token in extract_domain_like(m.content or ""):
            host = token.split("/")[0].lower()
            if "." in host:
                tld = "." + host.split(".")[-1]
                tld_counts[tld] += 1

    # False positive/negative approximations based on feedback vs verdict
    fp = 0
    fn = 0
    verdict_by_message = {m.id: m.verdict for m in messages}
    for f in feedback:
        verdict = verdict_by_message.get(f.message_id)
        if f.is_phishing == 1 and verdict == "ALLOW":
            fn += 1
        if f.is_phishing == 0 and verdict in {"FLAG", "BLOCK"}:
            fp += 1

    # Serialize trends in date order
    dates = sorted(set(list(verdict_trend.keys()) + list(feedback_trend.keys())))
    verdict_series = [
        {"date": d, **verdict_trend[d]} for d in dates
    ]
    feedback_series = [
        {"date": d, **feedback_trend[d]} for d in dates
    ]

    return {
        "verdict_trend": verdict_series,
        "feedback_trend": feedback_series,
        "block_reasons": dict(block_reasons),
        "block_sources": dict(block_sources),
        "risk_histogram": buckets,
        "tld_distribution": dict(tld_counts),
        "errors": {"false_positive": fp, "false_negative": fn}
    }


@router.get("/trusted")
def list_trusted(db: Session = Depends(get_db)):
    items = db.query(TrustedSender).order_by(TrustedSender.created_at.desc()).all()
    return jsonable_encoder(items)


@router.post("/trusted")
def add_trusted(payload: dict, db: Session = Depends(get_db)):
    value = (payload.get("value") or "").strip().lower()
    type_value = (payload.get("type") or "").strip().lower()
    if not value or type_value not in {"email", "phone", "domain"}:
        raise HTTPException(status_code=400, detail="value and valid type required")
    item = TrustedSender(value=value, type=type_value, scope="admin")
    db.add(item)
    db.commit()
    db.refresh(item)
    return jsonable_encoder(item)

def _parse_upload(file: UploadFile) -> list[str]:
    name = (file.filename or "").lower()
    content = file.file.read()
    values: list[str] = []
    if name.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(StringIO(text))
        for row in reader:
            if not row:
                continue
            cell = str(row[0]).strip()
            if not cell:
                continue
            if cell.lower() in {"value", "email", "domain", "phone", "url"}:
                continue
            values.append(cell)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        wb = load_workbook(filename=BytesIO(content), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            cell = str(row[0]).strip() if row[0] is not None else ""
            if not cell:
                continue
            if cell.lower() in {"value", "email", "domain", "phone", "url"}:
                continue
            values.append(cell)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    return values


@router.post("/trusted/bulk")
def add_trusted_bulk(payload: dict, db: Session = Depends(get_db)):
    values = payload.get("values") or []
    type_value = (payload.get("type") or "").strip().lower()
    if not values or type_value not in {"email", "phone", "domain"}:
        raise HTTPException(status_code=400, detail="values and valid type required")
    added = 0
    for value in values:
        val = str(value).strip().lower()
        if not val:
            continue
        exists = db.query(TrustedSender).filter(
            TrustedSender.value == val,
            TrustedSender.scope == "admin"
        ).first()
        if exists:
            continue
        db.add(TrustedSender(value=val, type=type_value, scope="admin"))
        added += 1
    db.commit()
    return {"status": "ok", "added": added}


@router.post("/trusted/upload")
def add_trusted_upload(type: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    type_value = (type or "").strip().lower()
    if type_value not in {"email", "phone", "domain"}:
        raise HTTPException(status_code=400, detail="valid type required")
    values = _parse_upload(file)
    added = 0
    for value in values:
        val = str(value).strip().lower()
        if not val:
            continue
        exists = db.query(TrustedSender).filter(
            TrustedSender.value == val,
            TrustedSender.scope == "admin"
        ).first()
        if exists:
            continue
        db.add(TrustedSender(value=val, type=type_value, scope="admin"))
        added += 1
    db.commit()
    return {"status": "ok", "added": added}


@router.delete("/trusted/{item_id}")
def delete_trusted(item_id: int, db: Session = Depends(get_db)):
    item = db.query(TrustedSender).filter(TrustedSender.id == item_id, TrustedSender.scope == "admin").first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(item)
    db.commit()
    return {"status": "deleted"}


@router.get("/blocked")
def list_blocked(db: Session = Depends(get_db)):
    items = db.query(BlockedEntity).order_by(BlockedEntity.created_at.desc()).all()
    return jsonable_encoder(items)


@router.post("/blocked")
def add_blocked(payload: dict, db: Session = Depends(get_db)):
    value = (payload.get("value") or "").strip().lower()
    type_value = (payload.get("type") or "").strip().lower()
    if not value or type_value not in {"email", "phone", "domain", "url"}:
        raise HTTPException(status_code=400, detail="value and valid type required")
    item = BlockedEntity(value=value, type=type_value, scope="admin")
    db.add(item)
    db.commit()
    db.refresh(item)
    return jsonable_encoder(item)


@router.post("/blocked/bulk")
def add_blocked_bulk(payload: dict, db: Session = Depends(get_db)):
    values = payload.get("values") or []
    type_value = (payload.get("type") or "").strip().lower()
    if not values or type_value not in {"email", "phone", "domain", "url"}:
        raise HTTPException(status_code=400, detail="values and valid type required")
    added = 0
    for value in values:
        val = str(value).strip().lower()
        if not val:
            continue
        exists = db.query(BlockedEntity).filter(
            BlockedEntity.value == val,
            BlockedEntity.scope == "admin"
        ).first()
        if exists:
            continue
        db.add(BlockedEntity(value=val, type=type_value, scope="admin"))
        added += 1
    db.commit()
    return {"status": "ok", "added": added}


@router.post("/blocked/upload")
def add_blocked_upload(type: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    type_value = (type or "").strip().lower()
    if type_value not in {"email", "phone", "domain", "url"}:
        raise HTTPException(status_code=400, detail="valid type required")
    values = _parse_upload(file)
    added = 0
    for value in values:
        val = str(value).strip().lower()
        if not val:
            continue
        exists = db.query(BlockedEntity).filter(
            BlockedEntity.value == val,
            BlockedEntity.scope == "admin"
        ).first()
        if exists:
            continue
        db.add(BlockedEntity(value=val, type=type_value, scope="admin"))
        added += 1
    db.commit()
    return {"status": "ok", "added": added}


@router.delete("/blocked/{item_id}")
def delete_blocked(item_id: int, db: Session = Depends(get_db)):
    item = db.query(BlockedEntity).filter(BlockedEntity.id == item_id, BlockedEntity.scope == "admin").first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(item)
    db.commit()
    return {"status": "deleted"}

@router.post("/retrain")
def retrain():
    allowed, reason = can_retrain()
    if not allowed:
        return {
            "status": "blocked",
            "reason": reason
        }

    try:
        result = trigger_retrain()
        if isinstance(result, dict):
            LAST_TRAINING_SUMMARY.clear()
            LAST_TRAINING_SUMMARY.update(result)
        return {
            "status": "retraining started",
            "detail": result
        }
    except HTTPException as exc:
        return {
            "status": "failed",
            "reason": exc.detail
        }


@router.get("/training-summary")
def training_summary():
    return LAST_TRAINING_SUMMARY or {"status": "no_training_run"}


@router.get("/settings/slack-channel")
def get_slack_channel(db: Session = Depends(get_db)):
    item = db.query(AppSetting).filter(AppSetting.key == "slack_alert_channel").first()
    return {"channel": item.value if item else ""}


@router.post("/settings/slack-channel")
def set_slack_channel(payload: dict, db: Session = Depends(get_db)):
    channel = str(payload.get("channel", "")).strip()
    item = db.query(AppSetting).filter(AppSetting.key == "slack_alert_channel").first()
    if item:
        item.value = channel
    else:
        db.add(AppSetting(key="slack_alert_channel", value=channel))
    db.commit()
    return {"status": "ok", "channel": channel}


@router.get("/settings/slack-channels")
def list_slack_channels():
    token = os.getenv("SLACK_BOT_TOKEN", "")
    if not token:
        raise HTTPException(status_code=400, detail="Slack bot token not configured")
    channels = []
    cursor = None
    for _ in range(5):
        params = {"limit": 200, "exclude_archived": True}
        if cursor:
            params["cursor"] = cursor
        resp = requests.get(
            "https://slack.com/api/conversations.list",
            headers={"Authorization": f"Bearer {token}"},
            params=params,
            timeout=8
        )
        data = resp.json()
        if not data.get("ok"):
            raise HTTPException(status_code=400, detail=data.get("error", "Slack API error"))
        for ch in data.get("channels", []):
            channels.append({"id": ch.get("id"), "name": ch.get("name")})
        cursor = data.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break
    return {"channels": channels}
